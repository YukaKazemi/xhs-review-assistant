# xhs_user_spider_ai.py (v5.6 - 基于您的 v5.5 修改)
# - 【核心修复 I - 效率与稳定】: 彻底重构 save_to_fetched_list 和后台Excel写入逻辑。
#   1. Redis更新变为【即时同步】，确保内存数据100%准确。
#   2. Excel写入任务交由【后台线程池】使用【Pandas】处理，效率和稳定性远超旧版 openpyxl 逐行加载模式，大幅降低并发冲突。
#   3. 新增 _perform_save_to_excel 内部函数，专职负责带锁的、高效的Excel追加操作。
# - 【核心修复 II - 关键日志】: 在 save_to_fetched_list 中，为“ID已存在但用户名是新的”这一关键场景添加了明确的日志输出，便于追踪内存更新。
# - 【核心修复 III - 健壮性】: 重构 load_fetched_list_to_redis 函数，改用Pandas并增加列名兼容和错误处理，确保服务启动时数据加载的可靠性。
# - 【架构保留】: 您原有的AI模型、Redis、任务队列、WAL、所有API路由等高级功能全部保留，本次为外科手术式修复。

# -*- coding: utf-8 -*-
import os
import io
import json
import queue
import atexit
import signal
import hashlib
import logging
import threading
import re
import platform
import subprocess
import time
from datetime import datetime, timedelta
from typing import Optional, List, Dict

from concurrent.futures import ThreadPoolExecutor

import requests
import pandas as pd
from PIL import Image
import redis
from flask import Flask, request, jsonify, Response, send_file, render_template_string
from flask_cors import CORS
from werkzeug.serving import WSGIRequestHandler
from filelock import FileLock, Timeout
import portalocker # 使用更通用的文件锁

import torch
import joblib
import numpy as np
import torch.nn as nn
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook, load_workbook

# ========== 基础配置 (来自您的 v5.5) ==========
FLASK_PORT = 5001
APPROVED_EXCEL_PATH = "已通过数据.xlsx"
MANUAL_REVIEW_EXCEL_PATH = "待复审数据.xlsx"
NEW_TRAINING_DATA_EXCEL = "待训练数据.xlsx"
DELTA_PREFIX = "工作成果_"
# 【修改】将旧文件名作为常量，方便兼容
FETCHED_USER_LIST_PATH = "已爬取用户名.xlsx" 
# 【新】为分离后的文件定义新名称
FETCHED_USERNAMES_FILE = '已爬取用户名_v2.xlsx'
FETCHED_USERIDS_FILE = '已爬取用户ID_v2.xlsx'

IMAGES_ROOT = os.path.join("data", "images")
IMAGE_MAX_SIDE = 768; IMAGE_FORMAT = "WEBP"; IMAGE_QUALITY = 90
REDIS_HOST = "localhost"; REDIS_PORT = 6379; REDIS_DB = 0
BATCH_FLUSH_ROWS = 20; BATCH_FLUSH_SEC  = 10.0; MAX_QUEUE_SIZE = 2000
IMG_MAX_WORKERS  = 6; HTTP_TIMEOUT = (3, 8)
WAL_DIR   = os.path.join("data", "wal_final"); WAL_FILE  = os.path.join(WAL_DIR, "mark_data.jsonl")

# ========== Redis Keys (来自您的 v5.5) ==========
AI_ENABLED_KEY = "ai:enabled"
SAVE_HISTORY_ENABLED_KEY = "save_history:enabled"
WAL_DONE_SET = "wal:done_final"
USERNAMES_SET_KEY = "usernames_set"
USERIDS_SET_KEY = "userids_set"
CUR_APPROVED = "export_cursor:approved"
CUR_REJECTED = "export_cursor:rejected"

# ========== 锁 & 线程安全 (来自您的 v5.5) ==========
mark_data_lock = threading.Lock() 

# ========== 邮箱提取功能 (来自您的 v5.5) ==========
EMAIL_MAPPING_DICT = { "艾特": "@", " at ": "@", " at": "@", "at ": "@", "©": "@", "®": "@", "🍥": "@", "＠": "@", "(at)": "@", "[at]": "@", "At": "@", "A T": "@", "🐧": "qq", "qq号": "qq", "扣扣": "qq", "扣": "q", "q ": "q", "企鹅": "qq", "球球": "qq", "163": "163", "一六三": "163", "一二六": "126", "126": "126", "新浪": "sina", "谷歌": "gmail", "outlook": "outlook", "hotmail": "hotmail", "点": ".", "dian": ".", " dot ": ".", " dot": ".", "dot ": ".", "。": ".", "·": ".", "丶": ".", " . ": ".", "com": "com", "康姆": "com", "点com": ".com", "cn": "cn", "点cn": ".cn", "零": "0", "一": "1", "二": "2", "三": "3", "四": "4", "五": "5", "六": "6", "七": "7", "八": "8", "九": "9", }
def extract_and_normalize_email(text: str) -> Optional[str]:
    if not text: return None
    normalized_text = text.lower()
    for non_standard, standard in EMAIL_MAPPING_DICT.items(): normalized_text = normalized_text.replace(non_standard, standard)
    normalized_text = re.sub(r'[\s:：,，\(\)\[\]]', '', normalized_text)
    match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', normalized_text)
    if match: return match.group(0)
    qq_match = re.search(r'([1-9][0-9]{5,10})qq\.com', normalized_text)
    if qq_match: return f"{qq_match.group(1)}@qq.com"
    return None

# ========== AI 模型加载 (来自您的 v5.5) ==========
MODEL_PATH = 'blogger_classifier_model.pth'
SCALER_PATH = 'scaler.joblib'
TEXT_MODEL_NAME = r'F:\AI_Model_Project\huggingface_cache\moka-ai_m3e-base'
IMAGE_MODEL_NAME = 'sentence-transformers/clip-ViT-B-32'
NUM_NOTES_TO_PROCESS = 20
AI_REJECT_THRESHOLD = 0.3
AI_ACCEPT_THRESHOLD = 0.55
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'
os.environ['HF_HOME'] = './huggingface_cache'
print("🤖 正在初始化本地AI环境...")
AI_ENABLED_BY_FILE = False
try:
    if not os.path.exists(MODEL_PATH) or not os.path.exists(SCALER_PATH): raise FileNotFoundError(f"模型({MODEL_PATH})或标准化文件({SCALER_PATH})不存在。")
    AI_DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"   - 使用设备: {AI_DEVICE.upper()}")
    print(f"   - 正在加载分类器模型: {MODEL_PATH}")
    model_state_dict = torch.load(MODEL_PATH, map_location=AI_DEVICE, weights_only=True)
    input_dim = model_state_dict['network.0.weight'].shape[1]
    class BloggerClassifier(nn.Module):
        def __init__(self, input_features):
            super(BloggerClassifier, self).__init__(); self.network = nn.Sequential(nn.Linear(input_features, 512), nn.BatchNorm1d(512), nn.ReLU(), nn.Dropout(0.5), nn.Linear(512, 256), nn.BatchNorm1d(256), nn.ReLU(), nn.Dropout(0.5), nn.Linear(256, 128), nn.BatchNorm1d(128), nn.ReLU(), nn.Dropout(0.4), nn.Linear(128, 1), nn.Sigmoid())
        def forward(self, x): return self.network(x)
    AI_MODEL = BloggerClassifier(input_features=input_dim); AI_MODEL.load_state_dict(model_state_dict); AI_MODEL.to(AI_DEVICE); AI_MODEL.eval()
    print("   - 分类器加载成功!")
    print(f"   - 正在加载Scaler: {SCALER_PATH}"); AI_SCALER = joblib.load(SCALER_PATH); print("   - Scaler加载成功!")
    print(f"   - 正在加载文本模型: {TEXT_MODEL_NAME} ..."); TEXT_EMBEDDING_MODEL = SentenceTransformer(TEXT_MODEL_NAME, device=AI_DEVICE); print("   - 文本模型加载成功!")
    print(f"   - 正在加载图像模型: {IMAGE_MODEL_NAME} ..."); IMAGE_EMBEDDING_MODEL = SentenceTransformer(IMAGE_MODEL_NAME, device=AI_DEVICE); print("   - 图像模型加载成功!")
    print("\n✅ AI模型全部加载成功！服务准备就绪。")
    AI_ENABLED_BY_FILE = True
except Exception as e:
    print(f"\n❌ AI模型加载失败: {e}\n   - 服务将以无AI模式运行。")

# ========== Flask / Redis / 工具函数 (来自您的 v5.5) ==========
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
CORS(app)
WSGIRequestHandler.protocol_version = "HTTP/1.1"
logging.getLogger('werkzeug').disabled = True
# 【修改】使用更规范的日志格式
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] [%(levelname)s] %(message)s', datefmt='%H:%M:%S')

r = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=REDIS_DB, decode_responses=True)
http = requests.Session()
job_q = queue.Queue(maxsize=MAX_QUEUE_SIZE)
img_pool = ThreadPoolExecutor(max_workers=IMG_MAX_WORKERS)
# 【新】为Excel写入操作创建一个专用的后台线程池
excel_writer_pool = ThreadPoolExecutor(max_workers=1, thread_name_prefix='ExcelWriter')

shutdown_event = threading.Event()
os.makedirs(IMAGES_ROOT, exist_ok=True)
os.makedirs(WAL_DIR, exist_ok=True)

# ... (保留您v5.5的工具函数)
def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def parse_dt(s: str) -> Optional[datetime]:
    if not s: return None
    for fmt in ("%Y-m-d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-m-%d", "%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None
def parse_count_chinese(s: str) -> int:
    s = str(s); s = s.translate(str.maketrans({**{chr(ord('０')+i): str(i) for i in range(10)}, '．': '.', '，': ','})).replace(",", "").rstrip("+").strip()
    if s.endswith("万"): return int(round(float(s[:-1]) * 10000)) if s[:-1] else 0
    return int(float(s)) if s and s.replace('.', '', 1).isdigit() else 0
def get_statistical_date_str(dt: datetime = None) -> str:
    dt = dt or datetime.now(); return (dt - timedelta(days=1) if dt.hour < 4 else dt).strftime('%Y-%m-%d')
def update_and_print_daily_stats(status: str):
    try:
        date_str = get_statistical_date_str(); approved_key, rejected_key = f"daily_stats:approved:{date_str}", f"daily_stats:rejected:{date_str}"
        (r.incr(approved_key) if status == "符合" else (r.incr(rejected_key) if status == "不符合" or status == "人工审核" else None))
        r.expire(approved_key, 48 * 3600); r.expire(rejected_key, 48 * 3600)
        approved_count, rejected_count = int(r.get(approved_key) or 0), int(r.get(rejected_key) or 0)
        logging.info(f"📊 今日统计 (4AM-4AM): 符合 {approved_count} | 其他 {rejected_count}")
    except Exception as e: logging.error(f"❌ 更新每日统计失败: {e}")

INFO_SHEET = "博主信息"; NOTES_SHEET = "博主笔记"
APPROVED_COLS = ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "个人简介", "粉丝数", "总点赞", "标记时间"]
TRAINING_INFO_COLS = ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "审核状态", "AI预测概率", "AI预测状态", "个人简介", "粉丝数", "总点赞", "标记时间"]
TRAINING_NOTES_COLS = ["小红书号", "笔记序号", "笔记标题", "笔记点赞数", "笔记封面路径", "标记时间"]
REVIEW_COLS = ["URL", "小红书号", "用户名", "标记时间"]

# 【修改】保留您的 safe_write_with_lock 框架，但内部调用会改变
def safe_write_with_lock(xlsx_path: str, writer_func):
    lock_path = xlsx_path + ".lock"
    filename = os.path.basename(xlsx_path)
    try:
        with FileLock(lock_path, timeout=5):
            return writer_func(xlsx_path)
    except Timeout:
        logging.error(f"❌❌❌ 严重错误: 获取文件锁超时！'{filename}' 可能正被其他程序(如WPS/Office)占用。")
        raise IOError(f"获取文件锁超时: '{filename}' 可能被占用。")
    except Exception as e:
        logging.error(f"❌❌❌ 严重错误: 写入 '{filename}' 时发生未知错误: {e}")
        raise IOError(f"写入 '{filename}' 时出错: {e}")

# ... (保留您v5.5的 download_and_convert_image 函数)
def download_and_convert_image(url: str, userid: str) -> str:
    if not url or not userid: return ""
    try:
        resp = http.get(url, timeout=HTTP_TIMEOUT, stream=True)
        if resp.status_code != 200: return ""
        raw = resp.content; h = hashlib.sha1(raw).hexdigest(); subdir = os.path.join(IMAGES_ROOT, userid, h[:2]); os.makedirs(subdir, exist_ok=True)
        out_path = os.path.join(subdir, f"{h}.webp")
        if os.path.exists(out_path): return os.path.relpath(out_path).replace("\\", "/")
        im = Image.open(io.BytesIO(raw)).convert("RGB"); w, hgt = im.size
        if max(w, hgt) > IMAGE_MAX_SIDE: ratio = IMAGE_MAX_SIDE / max(w, hgt); im = im.resize((int(w * ratio), int(hgt * ratio)), Image.LANCZOS)
        im.save(out_path, format=IMAGE_FORMAT, quality=IMAGE_QUALITY, method=6)
        return os.path.relpath(out_path).replace("\\", "/")
    except Exception: return ""

# ========== 后台任务处理 (核心修改区) ==========

# 【新】高效的后台Excel写入函数，使用Pandas
def _perform_save_to_excel(filepath: str, data_list: List[str], column_name: str):
    lock_path = filepath + ".lock"
    try:
        with portalocker.Lock(lock_path, 'a+', timeout=10):
            existing_df = pd.DataFrame()
            if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
                try:
                    existing_df = pd.read_excel(filepath)
                except Exception as e:
                    logging.warning(f"读取现有 {filepath} 失败，将创建新文件。错误: {e}")

            new_data_df = pd.DataFrame(data_list, columns=[column_name])
            combined_df = pd.concat([existing_df, new_data_df], ignore_index=True).drop_duplicates(subset=[column_name], keep='last')
            combined_df.to_excel(filepath, index=False)
            logging.info(f"💾 (后台)成功将 {len(data_list)} 条记录写入 {os.path.basename(filepath)}")
    except portalocker.LockException:
        logging.error(f"❌❌❌ Excel写入失败: 获取文件锁超时! '{filepath}' 可能被占用。")
    except Exception as e:
        logging.error(f"❌❌❌ 在后台写入 {filepath} 时发生严重错误: {e}")

# 【重构】新的 save_to_fetched_list 函数
def save_to_fetched_list(username: Optional[str] = None, userid: Optional[str] = None):
    """
    v5.6核心函数：负责将用户/ID同步到Redis和Excel。
    1. 立即更新Redis (内存)。
    2. 将持久化到Excel的任务异步提交到后台。
    """
    username = username.strip() if username else None
    userid = str(userid).strip() if userid else None

    # 1. 即时更新Redis
    if username:
        is_uname_new = r.sadd(USERNAMES_SET_KEY, username) == 1
        if is_uname_new:
            excel_writer_pool.submit(_perform_save_to_excel, FETCHED_USERNAMES_FILE, [username], "用户名")
    if userid:
        is_uid_new = r.sadd(USERIDS_SET_KEY, userid) == 1
        if is_uid_new:
            excel_writer_pool.submit(_perform_save_to_excel, FETCHED_USERIDS_FILE, [userid], "小红书号")
    
    # 2. 【关键日志】回应您的核心问题
    if username and userid:
        was_id_present = not is_uid_new
        if was_id_present and is_uname_new:
            logging.info(f"✨ 关键场景: ID '{userid}' 已存在，新增关联用户名 '{username}' 到Redis和Excel。")


class ApprovedBatcher: # 保留您的实现
    def __init__(self):
        self.lock = threading.Lock(); self.info_rows: List[Dict] = []; self.last_flush_time = time.time()
    def add(self, info_row: Dict):
        with self.lock: self.info_rows.append(info_row)
    def flush(self, force: bool = False):
        with self.lock:
            now = time.time()
            if not force and (not self.info_rows or (len(self.info_rows) < BATCH_FLUSH_ROWS and now - self.last_flush_time < BATCH_FLUSH_SEC)): return
            rows_to_write = list(self.info_rows); self.info_rows = self.info_rows[len(rows_to_write):]
        if not rows_to_write: return
        
        def writer(path):
            if not os.path.exists(path):
                wb = Workbook(); ws = wb.active; ws.title = INFO_SHEET; ws.append(APPROVED_COLS); wb.save(path)
            wb = load_workbook(path); ws = wb.active
            for row_dict in rows_to_write: ws.append([row_dict.get(h, "") for h in APPROVED_COLS])
            wb.save(path)
        try:
            safe_write_with_lock(APPROVED_EXCEL_PATH, writer)
            logging.info(f"📦 (后台)批量写入 '已通过' {len(rows_to_write)} 条成功。")
        except Exception as e: logging.error(f"⚠️ (后台)批量写入 '已通过' 失败，将重试: {e}")
approved_batcher = ApprovedBatcher()

def save_for_approved(job: Dict):
    # 【核心修复】将 "邮箱": "email" 添加到 update 字典中
    info_row_map = {h: h for h in ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "个人简介", "粉丝数", "总点赞", "标记时间"]}
    info_row_map.update({
        "用户名": "username", 
        "小红书号": "userid", 
        "主页网址": "url", 
        "邮箱": "email",  # <-- 这就是缺失的关键映射！
        "个人简介": "bio", 
        "粉丝数": "followers", 
        "总点赞": "likes_total", 
        "标记时间": "timestamp", 
        "搜索词": "search_term"
    })
    info_data = {k: job.get(v, "") for k, v in info_row_map.items()}
    approved_batcher.add(info_data)

def save_for_review(job: Dict): # 保留您的实现
    row = [job.get("url", ""), job.get("userid", ""), job.get("username", ""), job.get("timestamp", "")]
    def writer(path):
        if not os.path.exists(path):
            wb = Workbook(); ws = wb.active; ws.append(REVIEW_COLS); wb.save(path)
        wb = load_workbook(path); wb.active.append(row); wb.save(path)
    safe_write_with_lock(MANUAL_REVIEW_EXCEL_PATH, writer)

# xhs_user_spider_ai.py -> 替换整个 save_for_training 函数
def save_for_training(job: Dict):
    # 【核心修复】将 "邮箱": "email" 添加到 update 字典中
    info_row_map = {h: h for h in TRAINING_INFO_COLS}
    info_row_map.update({
        "用户名": "username", 
        "小红书号": "userid", 
        "主页网址": "url", 
        "邮箱": "email", # <-- 这里也加上，确保万无一失
        "审核状态": "status", 
        "AI预测概率": "ai_prob", 
        "AI预测状态": "ai_decision", 
        "个人简介": "bio", 
        "粉丝数": "followers", 
        "总点赞": "likes_total", 
        "标记时间": "timestamp", 
        "搜索词": "search_term"
    })
    info_row = {k: job.get(v, "") for k, v in info_row_map.items()}
    
    # 后续笔记处理逻辑保持不变
    notes_rows = []
    userid = job.get("userid") or job.get("username") or "unknown"
    futures = {i: img_pool.submit(download_and_convert_image, n.get("cover_url", ""), userid) for i, n in enumerate(job.get("notes", [])[:NUM_NOTES_TO_PROCESS]) if n.get("cover_url")}
    covers = {i: fut.result() for i, fut in futures.items()}
    for i, n in enumerate(job.get("notes", [])[:NUM_NOTES_TO_PROCESS]): 
        notes_rows.append([job.get("userid", ""), i + 1, n.get("title", ""), parse_count_chinese(str(n.get("likes", "0"))), covers.get(i, ""), job.get("timestamp", "")])
    
    def writer(path):
        if not os.path.exists(path):
            wb = Workbook()
            ws_info = wb.active
            ws_info.title = INFO_SHEET
            ws_info.append(TRAINING_INFO_COLS)
            ws_notes = wb.create_sheet(NOTES_SHEET)
            ws_notes.append(TRAINING_NOTES_COLS)
            wb.save(path)
        wb = load_workbook(path)
        ws_info = wb[INFO_SHEET] if INFO_SHEET in wb.sheetnames else wb.create_sheet(INFO_SHEET, 0)
        ws_info.append([info_row.get(h, "") for h in TRAINING_INFO_COLS])
        if notes_rows:
            ws_notes = wb[NOTES_SHEET] if NOTES_SHEET in wb.sheetnames else wb.create_sheet(NOTES_SHEET, 1)
            for r in notes_rows: 
                ws_notes.append(r)
        wb.save(path)
    
    safe_write_with_lock(NEW_TRAINING_DATA_EXCEL, writer)

# 【修改】简化 process_job
def process_job(job: Dict):
    job_id = job.get("job_id")
    if job_id and r.sismember(WAL_DONE_SET, job_id): 
        return
    status = job.get("status")

    # 统一调用新的去重函数
    save_to_fetched_list(username=job.get("username"), userid=job.get("userid"))

    if status == "符合":
        try: save_for_approved(job)
        except Exception as e: logging.error(f"❌ [后台] 保存 '已通过' 数据失败: {e}")
    if status == "人工审核":
        try: save_for_review(job); logging.info(f"📋 (后台)已保存至待复审: {job.get('username') or job.get('userid')}")
        except Exception as e: logging.error(f"❌ [后台] 保存 '待复审' 数据失败: {e}")
        
    save_history_enabled = r.get(SAVE_HISTORY_ENABLED_KEY) == "1"
    if status == "符合" or (status == "不符合" and save_history_enabled):
        try: save_for_training(job)
        except Exception as e: logging.warning(f"⚠️ [后台] 写入训练数据失败: {e}")
        
    if job_id: 
        r.sadd(WAL_DONE_SET, job_id)

def consume_jobs(): # 保留您的实现
    logging.info("✅ 后台消费线程已启动...")
    while not shutdown_event.is_set():
        try:
            job = job_q.get(timeout=1)
            process_job(job)
            job_q.task_done()
        except queue.Empty:
            approved_batcher.flush()
            continue
        except Exception as e: logging.error(f"❌ 处理后台任务时发生严重错误: {e}"); time.sleep(2)

# ========== AI 预测核心函数 (来自您的 v5.5) ==========
def run_ai_prediction(blogger_data: Dict) -> (float, str, str):
    if not AI_ENABLED_BY_FILE: return 0.0, "error", "AI模型未启用或加载失败"
    try:
        text_dim=TEXT_EMBEDDING_MODEL.get_sentence_embedding_dimension()or 768;image_dim=IMAGE_EMBEDDING_MODEL.get_sentence_embedding_dimension()or 512;bio=str(blogger_data.get('bio',''));notes_data=blogger_data.get('notes',[]);found_email=extract_and_normalize_email(bio);has_email=1 if found_email else 0;followers=parse_count_chinese(str(blogger_data.get('followers',0)));total_likes=parse_count_chinese(str(blogger_data.get('likes_total',0)));avg_likes_per_fan=total_likes/(followers+1)if followers>0 else 0;s_ratio,d_ratio=0,0
        if notes_data:likes_list=[parse_count_chinese(str(n.get('likes','0')))for n in notes_data[:NUM_NOTES_TO_PROCESS]];s_ratio=sum(1 for l in likes_list if l<10)/len(likes_list)if likes_list else 0;d_ratio=sum(1 for l in likes_list if 10<=l<100)/len(likes_list)if likes_list else 0
        bio_vec=TEXT_EMBEDDING_MODEL.encode([bio],convert_to_tensor=True,device=AI_DEVICE);titles=[n.get('title','')for n in notes_data[:NUM_NOTES_TO_PROCESS]];title_vec=TEXT_EMBEDDING_MODEL.encode(titles,convert_to_tensor=True,device=AI_DEVICE).mean(axis=0,keepdim=True)if titles else torch.zeros((1,text_dim),device=AI_DEVICE);image_urls=[note.get('cover_url')for note in notes_data if note.get('cover_url')];image_vec=torch.zeros((1,image_dim),device=AI_DEVICE)
        if image_urls:
            pil_images=[]
            for url in image_urls[:NUM_NOTES_TO_PROCESS]:
                if url:
                    try:pil_images.append(Image.open(io.BytesIO(http.get(url,timeout=5).content)).convert("RGB"))
                    except:pass
            if pil_images:image_vec=IMAGE_EMBEDDING_MODEL.encode(pil_images,convert_to_tensor=True,device=AI_DEVICE).mean(axis=0,keepdim=True)
        numeric_features=np.array([[has_email,followers,total_likes,avg_likes_per_fan,s_ratio,d_ratio]],dtype=np.float32);full_feature_np=np.concatenate([numeric_features,bio_vec.cpu().numpy(),title_vec.cpu().numpy(),image_vec.cpu().numpy()],axis=1)
        if full_feature_np.shape[1]<len(AI_SCALER.feature_names_in_):padded_arr=np.zeros((1,len(AI_SCALER.feature_names_in_)));padded_arr[:,:full_feature_np.shape[1]]=full_feature_np;full_feature_np=padded_arr
        scaled_features=AI_SCALER.transform(pd.DataFrame(full_feature_np,columns=AI_SCALER.feature_names_in_))
        with torch.no_grad():probability=AI_MODEL(torch.tensor(scaled_features,dtype=torch.float32).to(AI_DEVICE)).item()
        if probability>AI_ACCEPT_THRESHOLD:decision="符合";reason=f"模型概率 {probability:.2f} > {AI_ACCEPT_THRESHOLD}"
        elif probability>=AI_REJECT_THRESHOLD:decision="人工审核";reason=f"模型概率 {probability:.2f} 在 [{AI_REJECT_THRESHOLD}, {AI_ACCEPT_THRESHOLD}] 之间"
        else:decision="不符合";reason=f"模型概率 {probability:.2f} < {AI_REJECT_THRESHOLD}"
        return round(probability,4),decision,reason
    except Exception as e:
        logging.error(f"❌ AI Prediction Error: {e}");return 0.0,"error",str(e)
# (找到 ai_extract_email_by_model 函数，并用下面的代码完全替换它)

def ai_extract_email_by_model(blogger_data: Dict) -> Optional[str]:
    """
    【V4 - 最终版AI邮箱提取】
    1. 【核心修正】直接在最原始的简介文本上操作，保留所有特殊字符和颜文字。
    2. 使用一个极其宽容的正则表达式来捕捉任何形式的'xxx@yyy'候选者，无论字符多怪异。
    3. AI模型通过在原始文本上进行“挖洞”假设检验，来判断哪个“怪异”的字符串是关键信息。
    4. 返回的是原始字符串，交由后续逻辑处理。
    """
    if not AI_ENABLED_BY_FILE:
        return None

    original_bio = blogger_data.get("bio", "")
    if not original_bio:
        return None

    try:
        # 1. 【核心修正】直接在原始bio上寻找候选者
        # 这个正则极其宽容：
        # - `[^\s@]+` 匹配任何非空白、非@的字符一次或多次 (用户名部分)
        # - `@{1}` 匹配一个@符号
        # - `[^\s@]+` 再次匹配任何非空白、非@的字符 (域名部分)
        # 它可以完美匹配 "📢517656306@𝓺𝓺.𝓬𝓸𝓶"
        candidates = re.findall(r'[^\s@]+@{1}[^\s@]+', original_bio)

        # 备用策略：如果找不到@符号，我们依然在“净化后”的文本里寻找长数字串
        if not candidates:
            normalized_bio_for_numeric = original_bio.lower()
            for non_standard, standard in EMAIL_MAPPING_DICT.items():
                normalized_bio_for_numeric = normalized_bio_for_numeric.replace(non_standard, standard)
            
            numeric_candidates = re.findall(r'[1-9][0-9]{4,}', normalized_bio_for_numeric)
            if numeric_candidates:
                longest_numeric = max(numeric_candidates, key=len)
                candidates.append(longest_numeric)

        if not candidates:
            logging.info("🤖 AI邮箱识别: 在原始文本中未找到任何候选联系方式。")
            return None

        # 2. 定义内联的AI预测函数 (它将在原始文本上工作)
        def get_prob_for_bio(temp_bio: str) -> float:
            temp_data = blogger_data.copy()
            temp_data['bio'] = temp_bio
            text_dim=TEXT_EMBEDDING_MODEL.get_sentence_embedding_dimension()or 768;image_dim=IMAGE_EMBEDDING_MODEL.get_sentence_embedding_dimension()or 512;has_email=1;followers=parse_count_chinese(str(temp_data.get('followers',0)));total_likes=parse_count_chinese(str(temp_data.get('likes_total',0)));avg_likes_per_fan=total_likes/(followers+1)if followers>0 else 0;s_ratio,d_ratio=0,0;notes_data=temp_data.get('notes',[]);
            if notes_data:likes_list=[parse_count_chinese(str(n.get('likes','0')))for n in notes_data[:NUM_NOTES_TO_PROCESS]];s_ratio=sum(1 for l in likes_list if l<10)/len(likes_list)if likes_list else 0;d_ratio=sum(1 for l in likes_list if 10<=l<100)/len(likes_list)if likes_list else 0
            bio_vec=TEXT_EMBEDDING_MODEL.encode([temp_data['bio']],convert_to_tensor=True,device=AI_DEVICE)
            titles=[n.get('title','')for n in notes_data[:NUM_NOTES_TO_PROCESS]];title_vec=TEXT_EMBEDDING_MODEL.encode(titles,convert_to_tensor=True,device=AI_DEVICE).mean(axis=0,keepdim=True)if titles else torch.zeros((1,text_dim),device=AI_DEVICE);image_urls=[note.get('cover_url')for note in notes_data if note.get('cover_url')];image_vec=torch.zeros((1,image_dim),device=AI_DEVICE)
            numeric_features=np.array([[has_email,followers,total_likes,avg_likes_per_fan,s_ratio,d_ratio]],dtype=np.float32);full_feature_np=np.concatenate([numeric_features,bio_vec.cpu().numpy(),title_vec.cpu().numpy(),image_vec.cpu().numpy()],axis=1)
            if full_feature_np.shape[1]<len(AI_SCALER.feature_names_in_):padded_arr=np.zeros((1,len(AI_SCALER.feature_names_in_)));padded_arr[:,:full_feature_np.shape[1]]=full_feature_np;full_feature_np=padded_arr
            scaled_features=AI_SCALER.transform(pd.DataFrame(full_feature_np,columns=AI_SCALER.feature_names_in_))
            with torch.no_grad():return AI_MODEL(torch.tensor(scaled_features,dtype=torch.float32).to(AI_DEVICE)).item()

        # 3. 让AI在最原始的文本上进行评估
        base_probability = get_prob_for_bio(original_bio)
        best_candidate = None
        max_prob_drop = -1

        for cand in set(candidates):
            bio_without_cand = original_bio.replace(cand, "")
            prob_without_cand = get_prob_for_bio(bio_without_cand)
            prob_drop = base_probability - prob_without_cand
            
            if prob_drop > max_prob_drop:
                max_prob_drop = prob_drop
                best_candidate = cand
        
        # 4. 智能决策与返回
        if max_prob_drop > 0.03:
            final_contact_info = best_candidate
            logging.info(f"🤖 AI邮箱识别: 找到关键信息 '{final_contact_info}' (移除后概率下降 {max_prob_drop:.2%})")
            
            # 【重要】返回的是原始字符串，我们相信后续的标准函数能处理它
            # 但我们可以在这里做一个最终的“翻译”尝试，返回一个更干净的版本
            normalized_result = final_contact_info.lower()
            for non_standard, standard in EMAIL_MAPPING_DICT.items():
                normalized_result = normalized_result.replace(non_standard, standard)

            # 如果翻译后是纯数字，补全qq.com
            if normalized_result.isdigit():
                return f"{normalized_result}@qq.com"

            # 否则返回翻译后的结果
            return normalized_result
        else:
            logging.info(f"🤖 AI邮箱识别: 未找到关键联系方式 (最大概率变化 {max_prob_drop:.2%})")
            return None

    except Exception as e:
        logging.error(f"❌ AI邮箱识别时出错: {e}")
        return None

# 【重构】新的健壮的数据加载函数
def load_fetched_list_to_redis():
    """v5.6: 使用Pandas从Excel加载数据到Redis，更健壮。"""
    
    # 1. 优先从新版独立文件加载
    if os.path.exists(FETCHED_USERNAMES_FILE):
        df_users = pd.read_excel(FETCHED_USERNAMES_FILE)
        if "用户名" in df_users.columns:
            users_to_add = df_users["用户名"].dropna().astype(str).tolist()
            if users_to_add: r.sadd(USERNAMES_SET_KEY, *users_to_add)
            logging.info(f"✅ 从 {FETCHED_USERNAMES_FILE} 加载 {len(users_to_add)} 个用户名到Redis。")

    if os.path.exists(FETCHED_USERIDS_FILE):
        df_ids = pd.read_excel(FETCHED_USERIDS_FILE)
        if "小红书号" in df_ids.columns:
            ids_to_add = df_ids["小红书号"].dropna().astype(str).tolist()
            if ids_to_add: r.sadd(USERIDS_SET_KEY, *ids_to_add)
            logging.info(f"✅ 从 {FETCHED_USERIDS_FILE} 加载 {len(ids_to_add)} 个用户ID到Redis。")

    # 2. 兼容并迁移旧版合并文件
    if os.path.exists(FETCHED_USER_LIST_PATH):
        logging.warning(f"检测到旧版文件 '{FETCHED_USER_LIST_PATH}'，将进行一次性迁移。")
        try:
            df = pd.read_excel(FETCHED_USER_LIST_PATH, sheet_name=None)
            users_migrated, ids_migrated = 0, 0
            
            if "爬取用户名" in df:
                col_name = "爬取用户名" if "爬取用户名" in df["爬取用户名"].columns else ("用户名" if "用户名" in df["爬取用户名"].columns else None)
                if col_name:
                    users_to_add = df["爬取用户名"][col_name].dropna().astype(str).tolist()
                    if users_to_add:
                        r.sadd(USERNAMES_SET_KEY, *users_to_add)
                        _perform_save_to_excel(FETCHED_USERNAMES_FILE, users_to_add, "用户名")
                        users_migrated = len(users_to_add)

            if "小红书号" in df:
                if "小红书号" in df["小红书号"].columns:
                    ids_to_add = df["小红书号"]["小红书号"].dropna().astype(str).tolist()
                    if ids_to_add:
                        r.sadd(USERIDS_SET_KEY, *ids_to_add)
                        _perform_save_to_excel(FETCHED_USERIDS_FILE, ids_to_add, "小红书号")
                        ids_migrated = len(ids_to_add)
            
            logging.info(f"✅ 旧文件迁移完成: {users_migrated} 用户名, {ids_migrated} ID。建议迁移后删除旧文件。")
            # os.rename(FETCHED_USER_LIST_PATH, FETCHED_USER_LIST_PATH + ".bak") # 可选：自动重命名
        except Exception as e:
            logging.error(f"❌ 迁移旧文件 '{FETCHED_USER_LIST_PATH}' 失败: {e}")


# ========== API 路由 (保留您的 v5.5 结构) ==========
@app.route("/usernames", methods=["GET"])
def get_usernames(): return jsonify(list(r.smembers(USERNAMES_SET_KEY)))
@app.route("/userids", methods=["GET"])
def get_userids(): return jsonify(list(r.smembers(USERIDS_SET_KEY)))
# (在 @app.route("/mark_data", methods=["POST"]) 这一行的上方，粘贴下面的新函数)

def _process_mark_data(data: Dict):
    """
    【新】这是 mark_data 的核心逻辑处理函数，不依赖于Web请求上下文。
    它接收一个数据字典，并完成所有后续处理。
    """
    username, userid = (data.get("username") or "").strip(), (data.get("userid") or "").strip()
    final_decision = (data.get("status") or "").strip()

    if not final_decision:
        ai_prob, ai_decision, _ = run_ai_prediction(data)
        final_decision = ai_decision
        data['status'] = final_decision; data['ai_prob'] = ai_prob; data['ai_decision'] = ai_decision
    else:
        ai_prob, ai_decision, _ = run_ai_prediction(data)
        data['ai_prob'] = ai_prob; data['ai_decision'] = ai_decision

    if not final_decision or (not username and not userid):
        logging.warning(f"[_process_mark_data] 缺少关键参数，任务中止。Data: {data}")
        return {"status": "error", "message": "缺少关键参数"}

    # ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼ 核心修复点 ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼
    with mark_data_lock:
        is_uname_member = r.sismember(USERNAMES_SET_KEY, username) if username else False
        is_uid_member = r.sismember(USERIDS_SET_KEY, str(userid)) if userid else False
        is_rereview = data.get("is_rereview", False) # <-- 新增：获取前端标志

        # 修改逻辑：只有当它是重复的、不是人工审核、并且不是复审任务时，才判定为重复
        if (is_uname_member or is_uid_member) and final_decision != "人工审核" and not is_rereview:
            logging.warning(f"发现重复提交（非复审模式），已跳过: user=({username}|{userid})")
            return {"status": "duplicated"}

        # 对于“符合”或“不符合”的决定（无论是首次还是复审），都确保其存在于Redis中
        if final_decision in ["符合", "不符合"]:
            if username: r.sadd(USERNAMES_SET_KEY, username)
            if userid: r.sadd(USERIDS_SET_KEY, str(userid))
    # ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲ 核心修复点 ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲

    # V6 最终版邮箱处理逻辑
    final_email = ""
    bio_text = data.get("bio", "")
    frontend_email = data.get("email")
    if isinstance(frontend_email, str) and frontend_email.strip():
        final_email = frontend_email.strip(); logging.info("📧 邮箱来源: 前端手动填写。")
    if not final_email and bio_text:
        final_email = extract_and_normalize_email(bio_text) or "";
        if final_email: logging.info("📧 邮箱来源: 标准函数识别。")
    if not final_email and bio_text:
        final_email = ai_extract_email_by_model(data) or "";
        if final_email: logging.info("📧 邮箱来源: AI模型识别。")
    data['email'] = final_email

    ts = now_str()
    job_id = f"{userid or username}:{int(time.time()*1000)}:{final_decision}"
    job = {**data, "job_id": job_id, "timestamp": ts, "status": final_decision}
    
    try:
        with open(WAL_FILE, "a", encoding="utf-8") as f: f.write(json.dumps(job, ensure_ascii=False) + "\n")
        job_q.put(job, timeout=0.1)
    except queue.Full:
        logging.error("服务繁忙，后台任务队列已满！")
        return {"status": "error", "message": "服务繁忙，队列已满"}
    except Exception as e:
        logging.error(f"写入WAL或入队时出错: {e}")
        return {"status": "error", "message": "内部错误"}
    
    logging.info(f"    -> 任务已入队: {final_decision} (AI预测: {ai_decision}, P={ai_prob:.4f}), user=({username}|{userid})")
    update_and_print_daily_stats(final_decision)
    return {"status": "ok", "message": "任务已接收"}


@app.route("/mark_data", methods=["POST"])
def mark_data():
    data = request.get_json(silent=True) or {}
    result = _process_mark_data(data)
    
    status_code = 200
    if result.get("status") == "error":
        if "缺少关键参数" in result.get("message", ""): status_code = 400
        elif "队列已满" in result.get("message", ""): status_code = 503
        else: status_code = 500
        
    return jsonify(result), status_code

@app.route("/ai/decide", methods=["POST"])
def ai_decide():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"status": "error", "message": "No data provided"}), 400

    p, dec, why = run_ai_prediction(data)
    data['status'] = dec
    
    # 【核心修复】直接在后台线程中调用核心逻辑函数，不再通过HTTP请求绕圈
    threading.Thread(target=_process_mark_data, args=(data,)).start()
    
    return jsonify({"decision": dec, "reason": why, "p_base": p})
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"status": "error", "message": "No data provided"}), 400

    p, dec, why = run_ai_prediction(data)
    data['status'] = dec
    
    # 【核心修复】直接在后台线程中调用核心逻辑函数，不再通过HTTP请求绕圈
    threading.Thread(target=_process_mark_data, args=(data,)).start()
    
    return jsonify({"decision": dec, "reason": why, "p_base": p})
    data = request.get_json(silent=True) or {}
    result = _process_mark_data(data)
    
    status_code = 200
    if result.get("status") == "error":
        if "缺少关键参数" in result.get("message", ""): status_code = 400
        elif "队列已满" in result.get("message", ""): status_code = 503
        else: status_code = 500
        
    return jsonify(result), status_code
# ... (保留您 v5.5 的其他路由)
@app.route("/ai/settings", methods=["POST"])
def ai_settings(): enabled=bool(request.get_json(silent=True).get("enabled",False));r.set(AI_ENABLED_KEY,"1" if enabled else "0");logging.info(f"⚙️ AI自动审核已 {'开启' if enabled else '关闭'}");return jsonify({"status":"ok","ai_enabled":enabled})
@app.route("/settings/save_history", methods=["POST"])
def save_history_settings(): enabled=bool(request.get_json(silent=True).get("enabled",False));r.set(SAVE_HISTORY_ENABLED_KEY,"1" if enabled else "0");logging.info(f"⚙️ 保存“不符合”的训练数据已 {'开启' if enabled else '关闭'}");return jsonify({"status":"ok","save_history_enabled":enabled})
@app.route("/ai/suggest", methods=["POST"])
def ai_predict_only(): blogger_data = request.get_json(silent=True); p, dec, why = run_ai_prediction(blogger_data); return jsonify({"decision": dec, "reason": why, "p_base": p})
@app.route("/get_review_list", methods=["GET"])
def get_review_list():
    if not os.path.exists(MANUAL_REVIEW_EXCEL_PATH): 
        return jsonify([])
    try:
        def reader(path):
            wb = load_workbook(path); ws = wb.active
            if ws.max_row < 2: return [] 
            urls = [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True) if row and row[0]]
            if urls: ws.delete_rows(2, len(urls)); wb.save(path)
            return urls
        urls = safe_write_with_lock(MANUAL_REVIEW_EXCEL_PATH, reader) or []
        logging.info(f"✅ 提供了 {len(urls)} 个待复审URL，并已清空列表。")
        return jsonify(urls)
    except Exception as e: 
        logging.error(f"❌ 读取 '{MANUAL_REVIEW_EXCEL_PATH}' 失败: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500

def read_sheet_as_dicts(xlsx_path: str, sheet_name: str) -> List[Dict]:
    if not os.path.exists(xlsx_path): return []
    wb=load_workbook(xlsx_path,read_only=True);ws=wb[sheet_name]
    if ws.max_row<2:return[]
    rows=list(ws.iter_rows(values_only=True));headers=[str(h or "").strip()for h in rows[0]];return[dict(zip(headers,r))for r in rows[1:]]
def _filter_delta_rows(rows: List[Dict], min_dt: Optional[datetime]) -> List[Dict]:
    return [d for d in rows if(dt:=parse_dt(str(d.get("标记时间",""))))and dt>min_dt] if min_dt else rows
@app.route("/export_delta", methods=["GET"])
def export_delta():
    dataset=(request.args.get("dataset")or"approved").lower()
    curA_dt=parse_dt(r.get(CUR_APPROVED))if r.exists(CUR_APPROVED)else None
    wb=Workbook();wb.remove(wb.active)
    if dataset in("approved","both"):
        ws_info=wb.create_sheet("已通过-博主信息");ws_info.append(APPROVED_COLS)
        if os.path.exists(APPROVED_EXCEL_PATH):
            all_info=read_sheet_as_dicts(APPROVED_EXCEL_PATH,INFO_SHEET)
            for d in _filter_delta_rows(all_info,curA_dt):ws_info.append([d.get(k,"")for k in APPROVED_COLS])
    ts_name=datetime.now().strftime("%Y%m%d_%H%M%S");out_name=f"{DELTA_PREFIX}{ts_name}.xlsx";wb.save(out_name);nowS=now_str()
    if dataset in("approved","both"):r.set(CUR_APPROVED,nowS)
    logging.info(f"📤 增量导出完成: {out_name}, 游标已更新至 {nowS}");return send_file(out_name,as_attachment=True,download_name=out_name)
@app.route("/rebuild_sets", methods=["POST"])
def rebuild_sets():
    try:
        logging.info("🔁 开始从所有Excel文件重建Redis去重集合..."); r.delete(USERNAMES_SET_KEY); r.delete(USERIDS_SET_KEY)
        for path in [APPROVED_EXCEL_PATH, NEW_TRAINING_DATA_EXCEL, MANUAL_REVIEW_EXCEL_PATH]:
            if not os.path.exists(path): continue
            for row in read_sheet_as_dicts(path, INFO_SHEET if path != MANUAL_REVIEW_EXCEL_PATH else "Sheet"):
                if u := (row.get("用户名") or row.get("URL") or "").strip(): r.sadd(USERNAMES_SET_KEY, u)
                if i := (row.get("小红书号") or "").strip(): r.sadd(USERIDS_SET_KEY, str(i))
        load_fetched_list_to_redis()
        final_user_count,final_id_count=r.scard(USERNAMES_SET_KEY),r.scard(USERIDS_SET_KEY)
        logging.info(f"✅ Redis去重集合已重建完成: {final_user_count} 用户名, {final_id_count} 小红书号。")
        return jsonify({"status":"ok","usernames":final_user_count,"userids":final_id_count})
    except Exception as e:return jsonify({"status":"error","msg":str(e)}),500
@app.route('/dashboard', methods=['GET'])
def dashboard_page():
    try:
        with open('dashboard.html','r',encoding='utf-8')as f:return render_template_string(f.read())
    except FileNotFoundError:return"Error: dashboard.html not found.",404
@app.route("/dashboard_stats", methods=["GET"])
def dashboard_stats():
    try:
        today_str=get_statistical_date_str();approved_today=int(r.get(f"daily_stats:approved:{today_str}")or 0);rejected_today=int(r.get(f"daily_stats:rejected:{today_str}")or 0);yesterday_dt=datetime.now()-timedelta(days=1);yesterday_str=get_statistical_date_str(yesterday_dt);approved_yesterday=int(r.get(f"daily_stats:approved:{yesterday_str}")or 0);rejected_yesterday=int(r.get(f"daily_stats:rejected:{yesterday_str}")or 0);pending_review_count=0
        if os.path.exists(MANUAL_REVIEW_EXCEL_PATH):
            def reader(path): return max(0,load_workbook(path).active.max_row-1)
            pending_review_count = safe_write_with_lock(MANUAL_REVIEW_EXCEL_PATH, reader) or 0
        return jsonify({"today":{"approved":approved_today,"rejected":rejected_today,"total":approved_today+rejected_today,},"yesterday":{"total":approved_yesterday+rejected_yesterday,},"pending_review":pending_review_count})
    except Exception as e:return jsonify({"status":"error","msg":str(e)}),500
@app.route('/open_folder', methods=['POST'])
def open_folder_route():
    key=(request.get_json()or{}).get('key');FILE_MAP={"approved":APPROVED_EXCEL_PATH,"review":MANUAL_REVIEW_EXCEL_PATH,"training":NEW_TRAINING_DATA_EXCEL,"output":os.getcwd()};file_path=FILE_MAP.get(key)
    if not file_path:return jsonify({"status":"error","message":"Invalid file key"}),400
    try:
        folder_path=os.path.dirname(os.path.abspath(file_path))if os.path.isfile(file_path)else os.path.abspath(file_path)
        if not os.path.exists(folder_path):os.makedirs(folder_path,exist_ok=True)
        system=platform.system()
        if system=="Windows":os.startfile(folder_path)
        elif system=="Darwin":subprocess.run(["open",folder_path])
        else:subprocess.run(["xdg-open",folder_path])
        logging.info(f"📂 已请求打开文件夹: {folder_path}");return jsonify({"status":"ok","path":folder_path})
    except Exception as e:return jsonify({"status":"error","message":str(e)}),500
@app.route('/download_file', methods=['GET'])
def download_file_route():
    key=request.args.get('key');FILE_MAP={"approved":APPROVED_EXCEL_PATH,"review":MANUAL_REVIEW_EXCEL_PATH,"training":NEW_TRAINING_DATA_EXCEL};file_path=FILE_MAP.get(key)
    if not file_path:return"Invalid file key",400
    if not os.path.exists(file_path):
        if key == 'approved': wb=Workbook();ws=wb.active;ws.title=INFO_SHEET;ws.append(APPROVED_COLS);wb.save(file_path)
        elif key == 'review': wb=Workbook();ws=wb.active;ws.append(REVIEW_COLS);wb.save(file_path)
        elif key == 'training': wb=Workbook();ws_info=wb.active;ws_info.title=INFO_SHEET;ws_info.append(TRAINING_INFO_COLS);ws_notes=wb.create_sheet(NOTES_SHEET);ws_notes.append(TRAINING_NOTES_COLS);wb.save(file_path)
    return send_file(file_path,as_attachment=True)

# 【修改】/touch_user 路由，统一调用新函数
@app.route("/touch_user", methods=["POST"])
def touch_user():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    userid   = (data.get("userid") or "").strip()
    if not username and not userid:
        return jsonify({"status":"error","message":"missing username/userid"}), 400
    try:
        # 统一调用新的、高效的函数，它会处理Redis和后台Excel写入
        save_to_fetched_list(username=username, userid=userid)
        return jsonify({"status":"ok"})
    except Exception as e:
        logging.error(f"touch_user error: {e}")
        return jsonify({"status":"error","message":str(e)}), 500

# ========== 启动与退出 (来自您的 v5.5) ==========
def graceful_shutdown(*args, **kwargs):
    logging.info("\n⏹️ 正在优雅退出，请稍候...")
    shutdown_event.set()
    logging.info("   - 等待任务队列清空..."); job_q.join()
    logging.info("   - 正在强制刷新所有批处理数据到Excel..."); approved_batcher.flush(force=True)
    excel_writer_pool.shutdown(wait=True) # 等待Excel写入任务完成
    logging.info("   - 刷新完成。")
    logging.info("✅ 后台任务已安全处理完毕，服务退出。")
    os._exit(0)

if __name__ == "__main__":
    consumer_thread = threading.Thread(target=consume_jobs, daemon=True); consumer_thread.start()
    atexit.register(graceful_shutdown)
    signal.signal(signal.SIGINT, graceful_shutdown); signal.signal(signal.SIGTERM, graceful_shutdown)
    load_fetched_list_to_redis()
    logging.info(f"🚀 [v5.6 - 基于v5.5修复版] 服务启动：http://127.0.0.1:{FLASK_PORT}")
    logging.info(f"📊 数据面板请访问: http://localhost:{FLASK_PORT}/dashboard")
    from waitress import serve
    serve(app, host="0.0.0.0", port=FLASK_PORT, threads=16)

# 删除了您代码末尾重复定义的路由
