# recover_from_jsonl.py - JSONL数据恢复工具
#
# 使用方法:
# 1. 将此文件放置在与 xhs_user_spider_ai.py 相同的项目根目录下。
# 2. 确保 Redis 服务正在运行。
# 3. 【重要】关闭所有可能正在打开的Excel文件（如 已通过数据.xlsx 等）。
# 4. 在终端中运行此脚本: python recover_from_jsonl.py
# 5. 脚本会自动读取 mark_data.jsonl 并将未处理的数据恢复到对应的Excel文件中。

import os
import io
import json
import hashlib
import threading
import time
from datetime import datetime
from typing import Optional, List, Dict
from concurrent.futures import ThreadPoolExecutor

import requests
from PIL import Image
import redis
from openpyxl import Workbook, load_workbook

# ========== 1. 从主程序复制过来的基础配置 (保持完全一致) ==========
APPROVED_EXCEL_PATH = "已通过数据.xlsx"
# REJECTED_EXCEL_PATH = "未通过数据.xlsx"
# NEW_TRAINING_DATA_EXCEL = "待训练数据.xlsx"
# FETCHED_USER_LIST_PATH = "已爬取用户名.xlsx" 
IMAGES_ROOT = os.path.join("data", "images")
IMAGE_MAX_SIDE = 768; IMAGE_FORMAT = "WEBP"; IMAGE_QUALITY = 90
REDIS_HOST = "localhost"; REDIS_PORT = 6379; REDIS_DB = 0
IMG_MAX_WORKERS  = 6; HTTP_TIMEOUT = (3, 8)
WAL_DIR   = os.path.join("data", "wal_final"); WAL_FILE  = os.path.join(WAL_DIR, "mark_data.jsonl")

# Redis Keys
SAVE_HISTORY_ENABLED_KEY = "save_history:enabled"
WAL_DONE_SET = "wal:done_final"

# 锁 (虽然单线程运行恢复脚本不是必须，但保持函数签名一致性)
fetched_user_list_lock = threading.Lock()
training_data_excel_lock = threading.Lock()
excel_write_lock = threading.Lock() # 通用Excel写锁

# ========== 2. 从主程序复制过来的核心工具函数 ==========
r = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=REDIS_DB, decode_responses=True)
http = requests.Session()
img_pool = ThreadPoolExecutor(max_workers=IMG_MAX_WORKERS)
os.makedirs(IMAGES_ROOT, exist_ok=True)

def parse_count_chinese(s: str) -> int:
    s = str(s)
    if not s or s == "赞": return 0
    trans = str.maketrans({**{chr(ord('０')+i): str(i) for i in range(10)}, '．': '.', '，': ','})
    s = s.translate(trans).replace(",", "").rstrip("+").strip()
    if s.endswith("万"):
        try: return int(round(float(s[:-1]) * 10000))
        except: return 0
    try: return int(float(s))
    except: return 0

INFO_SHEET = "博主信息"; NOTES_SHEET = "博主笔记"
INFO_BASE_COLS = ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "审核状态"]
INFO_AI_COLS = ["AI预测概率", "AI预测状态"]
INFO_EXTRA_COLS = ["个人简介", "粉丝数", "总点赞", "标记时间"]
ALL_INFO_COLS = INFO_BASE_COLS + INFO_AI_COLS + INFO_EXTRA_COLS

def ensure_workbook(xlsx_path: str):
    if not os.path.exists(xlsx_path): 
        wb = Workbook()
        ws_info = wb.active
        ws_info.title = INFO_SHEET
        ws_info.append(ALL_INFO_COLS)
        ws_notes = wb.create_sheet(NOTES_SHEET)
        ws_notes.append(["小红书号", "笔记序号", "笔记标题", "笔记点赞数", "笔记封面路径", "标记时间"])
        wb.save(xlsx_path)

def excel_append_batch(xlsx_path: str, info_rows: List[Dict], notes_rows: List[List]):
    with excel_write_lock:
        ensure_workbook(xlsx_path)
        wb = load_workbook(xlsx_path)
        if info_rows: 
            ws = wb[INFO_SHEET]
            headers = [c.value for c in ws[1]]
            [ws.append([row_dict.get(h, "") for h in headers]) for row_dict in info_rows]
        if notes_rows: 
            ws = wb[NOTES_SHEET]
            [ws.append(row) for row in notes_rows]
        wb.save(xlsx_path)

def download_and_convert_image(url: str, userid: str) -> str:
    if not url or not userid: return ""
    try:
        resp = http.get(url, timeout=HTTP_TIMEOUT, stream=True)
        if resp.status_code != 200: return ""
        raw = resp.content; h = hashlib.sha1(raw).hexdigest(); subdir = os.path.join(IMAGES_ROOT, userid, h[:2]); os.makedirs(subdir, exist_ok=True)
        out_path = os.path.join(subdir, f"{h}.webp")
        if os.path.exists(out_path): return os.path.relpath(out_path).replace("\\", "/")
        im = Image.open(io.BytesIO(raw)).convert("RGB"); w, hgt = im.size
        if max(w, hgt) > IMAGE_MAX_SIDE: ratio = IMAGE_MAX_SIDE / max(w, hgt); im = im.resize((int(w * ratio), int(hgt * ratio)), Image.LANCZOS)
        im.save(out_path, format=IMAGE_FORMAT, quality=IMAGE_QUALITY, method=6)
        return os.path.relpath(out_path).replace("\\", "/")
    except Exception: return ""

# ========== 3. 从主程序复制过来的数据处理逻辑函数 ==========
def append_to_fetched_list(username: Optional[str], userid: Optional[str]):
    if not username and not userid: return
    with fetched_user_list_lock:
        try:
            sheet_name_user, sheet_name_id = "爬取用户名", "小红书号"
            if not os.path.exists(FETCHED_USER_LIST_PATH): wb = Workbook(); ws_user = wb.active; ws_user.title = sheet_name_user; ws_user.append([sheet_name_user]); ws_id = wb.create_sheet(sheet_name_id); ws_id.append([sheet_name_id]); wb.save(FETCHED_USER_LIST_PATH)
            wb = load_workbook(FETCHED_USER_LIST_PATH)
            if username: ws_user = wb[sheet_name_user] if sheet_name_user in wb.sheetnames else wb.create_sheet(sheet_name_user, 0); ws_user.append([username])
            if userid: ws_id = wb[sheet_name_id] if sheet_name_id in wb.sheetnames else wb.create_sheet(sheet_name_id, 1); ws_id.append([str(userid)])
            wb.save(FETCHED_USER_LIST_PATH)
        except Exception as e: print(f"❌ (恢复)追加到 '{FETCHED_USER_LIST_PATH}' 失败: {e}")

def append_to_training_excel(job: Dict):
    if not job or not job.get("status"): return
    with training_data_excel_lock:
        try:
            info_row, notes_rows = build_excel_rows(job)
            excel_append_batch(NEW_TRAINING_DATA_EXCEL, [info_row], notes_rows)
        except Exception as e:
            print(f"❌ (恢复)写入训练数据到 '{NEW_TRAINING_DATA_EXCEL}' 失败: {e}")

def build_excel_rows(job: Dict):
    info_row_map = {
        "用户名": "username", "小红书号": "userid", "主页网址": "url", "邮箱": "email", "搜索词": "search_term", 
        "审核状态": "status", "AI预测概率": "ai_prob", "AI预测状态": "ai_decision", "个人简介": "bio", "粉丝数": "followers", 
        "总点赞": "likes_total", "标记时间": "timestamp"
    }
    info_row = {k: job.get(v, "") for k, v in info_row_map.items()}
    notes_rows = []; userid = job.get("userid") or job.get("username") or "unknown"
    futures = {i: img_pool.submit(download_and_convert_image, n.get("cover_url", ""), userid) for i, n in enumerate(job.get("notes", [])[:20]) if n.get("cover_url")}
    covers = {i: fut.result() for i, fut in futures.items()}
    for i, n in enumerate(job.get("notes", [])[:20]): notes_rows.append([job.get("userid", ""), i + 1, n.get("title", ""), parse_count_chinese(str(n.get("likes", "0"))), covers.get(i, ""), job.get("timestamp", "")])
    return info_row, notes_rows

# ========== 4. 恢复脚本的主逻辑 ==========
if __name__ == "__main__":
    print("--- 开始从 mark_data.jsonl 恢复数据到 Excel ---")
    print("ℹ️ 这个脚本不需要加载庞大的AI模型，只进行数据处理。")
    
    if not os.path.exists(WAL_FILE):
        print(f"❌ 未找到日志文件 '{WAL_FILE}'，无需恢复。")
        exit()
        
    try:
        r.ping()
        print("✅ Redis 连接成功。")
    except redis.exceptions.ConnectionError as e:
        print(f"❌ Redis 连接失败: {e}")
        print("   请确保您的Redis服务正在 localhost:6379 上运行。")
        exit()

    total_lines = 0
    processed_count = 0
    skipped_count = 0
    error_count = 0
    
    # 读取所有job到内存，方便显示进度
    with open(WAL_FILE, 'r', encoding='utf-8') as f:
        jobs_to_process = f.readlines()
    
    total_lines = len(jobs_to_process)
    print(f"📄 日志文件中共找到 {total_lines} 条记录。")

    for i, line in enumerate(jobs_to_process):
        line = line.strip()
        if not line:
            continue

        progress = f"[{i+1}/{total_lines}]"

        try:
            job = json.loads(line)
        except json.JSONDecodeError:
            print(f"⚠️ {progress} 警告: 解析JSON失败，跳过损坏的行。")
            error_count += 1
            continue

        job_id = job.get("job_id")
        if not job_id:
            print(f"⚠️ {progress} 警告: 记录缺少 'job_id'，跳过。")
            error_count += 1
            continue

        # 核心逻辑：检查是否已处理过
        if r.sismember(WAL_DONE_SET, job_id):
            # print(f"⏭️ {progress} 已跳过 (已处理): {job_id}") # 日志太长，默认不打印
            skipped_count += 1
            continue
        
        print(f"⚙️ {progress} 正在处理新任务: {job.get('username') or job.get('userid')} ({job.get('status')})")
        
        try:
            # 1. 写入训练数据 (包含图片下载)
            append_to_training_excel(job)
            
            # 2. 追加到已爬取列表
            append_to_fetched_list(job.get("username"), job.get("userid"))
            
            # 3. 写入主数据文件 (已通过/未通过)
            save_history_enabled = r.get(SAVE_HISTORY_ENABLED_KEY) == "1"
            status = job.get("status")
            
            if status == "符合" or (status == "不符合" and save_history_enabled):
                # build_excel_rows 已经在 append_to_training_excel 中被调用过一次
                # 为了效率，我们直接复用那个逻辑，或者再调用一次以确保数据隔离
                info_row, notes_rows = build_excel_rows(job)
                
                xlsx_path = APPROVED_EXCEL_PATH if status == "符合" else REJECTED_EXCEL_PATH
                excel_append_batch(xlsx_path, [info_row], notes_rows)
            
            # 4. 标记任务已完成
            r.sadd(WAL_DONE_SET, job_id)
            processed_count += 1

        except Exception as e:
            print(f"❌ {progress} 处理任务 {job_id} 时发生严重错误: {e}")
            error_count += 1
            
    print("\n" + "="*40)
    print("✅ 数据恢复完成！")
    print("="*40)
    print(f"📄 总记录数: {total_lines}")
    print(f"⚙️ 成功处理: {processed_count} 条")
    print(f"⏭️ 跳过 (已存在): {skipped_count} 条")
    print(f"❌ 失败/错误: {error_count} 条")
    print("\n现在您可以检查您的Excel文件了。")

