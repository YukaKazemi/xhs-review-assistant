# xhs_user_spider_ai.py (v5.5 - 关键修复版)
# - 【基准】本代码严格基于您提供的 v5.4 版本进行修改。
# - 【核心修复 I】修正了 /mark_data 路由的去重逻辑。现在只有状态为“符合”或“不符合”的用户才会被添加到Redis去重集合中。这彻底解决了之前已存在（如“不符合”）的用户无法被再次提交为“人工审核”状态的BUG。
# - 【核心修复 II】修正了 /get_review_list 路由的功能。采用更稳健的“删除行”方式来清空表格，而不是“创建新文件”，避免了潜在的文件权限和数据丢失问题，确保“复审”按钮能正常工作。
# - 【保留架构】保留了 v5.4 版本的所有优良特性，包括线程安全的Redis写入、后台任务队列、四通道数据存储等。

# -*- coding: utf-8 -*-
import os
import io
import json
import queue
import atexit
import signal
import hashlib
import logging
import threading
import re
import platform
import subprocess
import time
from datetime import datetime, timedelta
from typing import Optional, List, Dict

from concurrent.futures import ThreadPoolExecutor

import requests
import pandas as pd
from PIL import Image
import redis
from flask import Flask, request, jsonify, Response, send_file, render_template_string
from flask_cors import CORS
from werkzeug.serving import WSGIRequestHandler
from filelock import FileLock, Timeout

import torch
import joblib
import numpy as np
import torch.nn as nn
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook, load_workbook

# ========== 基础配置 ==========
FLASK_PORT = 5001
APPROVED_EXCEL_PATH = "已通过数据.xlsx"
MANUAL_REVIEW_EXCEL_PATH = "待复审数据.xlsx"
NEW_TRAINING_DATA_EXCEL = "待训练数据.xlsx"
DELTA_PREFIX = "工作成果_"
FETCHED_USER_LIST_PATH = "已爬取用户名.xlsx" 
IMAGES_ROOT = os.path.join("data", "images")
IMAGE_MAX_SIDE = 768; IMAGE_FORMAT = "WEBP"; IMAGE_QUALITY = 90
REDIS_HOST = "localhost"; REDIS_PORT = 6379; REDIS_DB = 0
BATCH_FLUSH_ROWS = 20; BATCH_FLUSH_SEC  = 10.0; MAX_QUEUE_SIZE = 2000
IMG_MAX_WORKERS  = 6; HTTP_TIMEOUT = (3, 8)
WAL_DIR   = os.path.join("data", "wal_final"); WAL_FILE  = os.path.join(WAL_DIR, "mark_data.jsonl")

# ========== Redis Keys ==========
AI_ENABLED_KEY = "ai:enabled"
SAVE_HISTORY_ENABLED_KEY = "save_history:enabled"
WAL_DONE_SET = "wal:done_final"
USERNAMES_SET_KEY = "usernames_set"
USERIDS_SET_KEY = "userids_set"
CUR_APPROVED = "export_cursor:approved"
CUR_REJECTED = "export_cursor:rejected"

# ========== 锁 & 线程安全 ==========
mark_data_lock = threading.Lock() 

# ========== 邮箱提取功能 ==========
EMAIL_MAPPING_DICT = { "艾特": "@", " at ": "@", " at": "@", "at ": "@", "©": "@", "®": "@", "🍥": "@", "＠": "@", "(at)": "@", "[at]": "@", "At": "@", "A T": "@", "🐧": "qq", "qq号": "qq", "扣扣": "qq", "扣": "q", "q ": "q", "企鹅": "qq", "球球": "qq", "163": "163", "一六三": "163", "一二六": "126", "126": "126", "新浪": "sina", "谷歌": "gmail", "outlook": "outlook", "hotmail": "hotmail", "点": ".", "dian": ".", " dot ": ".", " dot": ".", "dot ": ".", "。": ".", "·": ".", "丶": ".", " . ": ".", "com": "com", "康姆": "com", "点com": ".com", "cn": "cn", "点cn": ".cn", "零": "0", "一": "1", "二": "2", "三": "3", "四": "4", "五": "5", "六": "6", "七": "7", "八": "8", "九": "9", }
def extract_and_normalize_email(text: str) -> Optional[str]:
    if not text: return None
    normalized_text = text.lower()
    for non_standard, standard in EMAIL_MAPPING_DICT.items(): normalized_text = normalized_text.replace(non_standard, standard)
    normalized_text = re.sub(r'[\s:：,，\(\)\[\]]', '', normalized_text)
    match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', normalized_text)
    if match: return match.group(0)
    qq_match = re.search(r'([1-9][0-9]{5,10})qq\.com', normalized_text)
    if qq_match: return f"{qq_match.group(1)}@qq.com"
    return None

# ========== AI 模型加载 ==========
MODEL_PATH = 'blogger_classifier_model.pth'
SCALER_PATH = 'scaler.joblib'
TEXT_MODEL_NAME = r'F:\AI_Model_Project\huggingface_cache\moka-ai_m3e-base'
IMAGE_MODEL_NAME = 'sentence-transformers/clip-ViT-B-32'
NUM_NOTES_TO_PROCESS = 20
AI_REJECT_THRESHOLD = 0.3
AI_ACCEPT_THRESHOLD = 0.55
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'
os.environ['HF_HOME'] = './huggingface_cache'
print("🤖 正在初始化本地AI环境...")
AI_ENABLED_BY_FILE = False
try:
    if not os.path.exists(MODEL_PATH) or not os.path.exists(SCALER_PATH): raise FileNotFoundError(f"模型({MODEL_PATH})或标准化文件({SCALER_PATH})不存在。")
    AI_DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"   - 使用设备: {AI_DEVICE.upper()}")
    print(f"   - 正在加载分类器模型: {MODEL_PATH}")
    model_state_dict = torch.load(MODEL_PATH, map_location=AI_DEVICE, weights_only=True)
    input_dim = model_state_dict['network.0.weight'].shape[1]
    class BloggerClassifier(nn.Module):
        def __init__(self, input_features):
            super(BloggerClassifier, self).__init__(); self.network = nn.Sequential(nn.Linear(input_features, 512), nn.BatchNorm1d(512), nn.ReLU(), nn.Dropout(0.5), nn.Linear(512, 256), nn.BatchNorm1d(256), nn.ReLU(), nn.Dropout(0.5), nn.Linear(256, 128), nn.BatchNorm1d(128), nn.ReLU(), nn.Dropout(0.4), nn.Linear(128, 1), nn.Sigmoid())
        def forward(self, x): return self.network(x)
    AI_MODEL = BloggerClassifier(input_features=input_dim); AI_MODEL.load_state_dict(model_state_dict); AI_MODEL.to(AI_DEVICE); AI_MODEL.eval()
    print("   - 分类器加载成功!")
    print(f"   - 正在加载Scaler: {SCALER_PATH}"); AI_SCALER = joblib.load(SCALER_PATH); print("   - Scaler加载成功!")
    print(f"   - 正在加载文本模型: {TEXT_MODEL_NAME} ..."); TEXT_EMBEDDING_MODEL = SentenceTransformer(TEXT_MODEL_NAME, device=AI_DEVICE); print("   - 文本模型加载成功!")
    print(f"   - 正在加载图像模型: {IMAGE_MODEL_NAME} ..."); IMAGE_EMBEDDING_MODEL = SentenceTransformer(IMAGE_MODEL_NAME, device=AI_DEVICE); print("   - 图像模型加载成功!")
    print("\n✅ AI模型全部加载成功！服务准备就绪。")
    AI_ENABLED_BY_FILE = True
except Exception as e:
    print(f"\n❌ AI模型加载失败: {e}\n   - 服务将以无AI模式运行。")

# ========== Flask / Redis / 工具函数 ==========
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
CORS(app)
WSGIRequestHandler.protocol_version = "HTTP/1.1"
logging.getLogger('werkzeug').disabled = True
r = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=REDIS_DB, decode_responses=True)
http = requests.Session()
job_q = queue.Queue(maxsize=MAX_QUEUE_SIZE)
img_pool = ThreadPoolExecutor(max_workers=IMG_MAX_WORKERS)
shutdown_event = threading.Event()
os.makedirs(IMAGES_ROOT, exist_ok=True)
os.makedirs(WAL_DIR, exist_ok=True)
def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def parse_dt(s: str) -> Optional[datetime]:
    if not s: return None
    for fmt in ("%Y-m-d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-m-%d", "%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None
def parse_count_chinese(s: str) -> int:
    s = str(s); s = s.translate(str.maketrans({**{chr(ord('０')+i): str(i) for i in range(10)}, '．': '.', '，': ','})).replace(",", "").rstrip("+").strip()
    if s.endswith("万"): return int(round(float(s[:-1]) * 10000)) if s[:-1] else 0
    return int(float(s)) if s and s.replace('.', '', 1).isdigit() else 0
def get_statistical_date_str(dt: datetime = None) -> str:
    dt = dt or datetime.now(); return (dt - timedelta(days=1) if dt.hour < 4 else dt).strftime('%Y-%m-%d')
def update_and_print_daily_stats(status: str):
    try:
        date_str = get_statistical_date_str(); approved_key, rejected_key = f"daily_stats:approved:{date_str}", f"daily_stats:rejected:{date_str}"
        (r.incr(approved_key) if status == "符合" else (r.incr(rejected_key) if status == "不符合" or status == "人工审核" else None))
        r.expire(approved_key, 48 * 3600); r.expire(rejected_key, 48 * 3600)
        approved_count, rejected_count = int(r.get(approved_key) or 0), int(r.get(rejected_key) or 0)
        print(f"📊 今日统计 (4AM-4AM): 符合 {approved_count} | 其他 {rejected_count}")
    except Exception as e: print(f"❌ 更新每日统计失败: {e}")

INFO_SHEET = "博主信息"; NOTES_SHEET = "博主笔记"
APPROVED_COLS = ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "个人简介", "粉丝数", "总点赞", "标记时间"]
TRAINING_INFO_COLS = ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "审核状态", "AI预测概率", "AI预测状态", "个人简介", "粉丝数", "总点赞", "标记时间"]
TRAINING_NOTES_COLS = ["小红书号", "笔记序号", "笔记标题", "笔记点赞数", "笔记封面路径", "标记时间"]
REVIEW_COLS = ["URL", "小红书号", "用户名", "标记时间"]
def safe_write_with_lock(xlsx_path: str, writer_func):
    lock_path = xlsx_path + ".lock"
    try:
        with FileLock(lock_path, timeout=1):
            return writer_func(xlsx_path)
    except Timeout: raise IOError(f"获取文件锁超时: '{os.path.basename(xlsx_path)}' 可能被占用。")
    except Exception as e: raise IOError(f"写入 '{os.path.basename(xlsx_path)}' 时出错: {e}")

def download_and_convert_image(url: str, userid: str) -> str:
    if not url or not userid: return ""
    try:
        resp = http.get(url, timeout=HTTP_TIMEOUT, stream=True)
        if resp.status_code != 200: return ""
        raw = resp.content; h = hashlib.sha1(raw).hexdigest(); subdir = os.path.join(IMAGES_ROOT, userid, h[:2]); os.makedirs(subdir, exist_ok=True)
        out_path = os.path.join(subdir, f"{h}.webp")
        if os.path.exists(out_path): return os.path.relpath(out_path).replace("\\", "/")
        im = Image.open(io.BytesIO(raw)).convert("RGB"); w, hgt = im.size
        if max(w, hgt) > IMAGE_MAX_SIDE: ratio = IMAGE_MAX_SIDE / max(w, hgt); im = im.resize((int(w * ratio), int(hgt * ratio)), Image.LANCZOS)
        im.save(out_path, format=IMAGE_FORMAT, quality=IMAGE_QUALITY, method=6)
        return os.path.relpath(out_path).replace("\\", "/")
    except Exception: return ""

# ========== 后台任务处理 ==========
class ApprovedBatcher:
    def __init__(self):
        self.lock = threading.Lock(); self.info_rows: List[Dict] = []; self.last_flush_time = time.time()
    def add(self, info_row: Dict):
        with self.lock: self.info_rows.append(info_row)
    def flush(self, force: bool = False):
        with self.lock:
            now = time.time()
            if not force and (not self.info_rows or (len(self.info_rows) < BATCH_FLUSH_ROWS and now - self.last_flush_time < BATCH_FLUSH_SEC)): return
            rows_to_write = list(self.info_rows); self.info_rows = self.info_rows[len(rows_to_write):]
        if not rows_to_write: return
        def writer(path):
            if not os.path.exists(path):
                wb = Workbook(); ws = wb.active; ws.title = INFO_SHEET; ws.append(APPROVED_COLS); wb.save(path)
            wb = load_workbook(path); ws = wb.active
            for row_dict in rows_to_write: ws.append([row_dict.get(h, "") for h in APPROVED_COLS])
            wb.save(path)
        try:
            safe_write_with_lock(APPROVED_EXCEL_PATH, writer)
            print(f"📦 (后台)批量写入 '已通过' {len(rows_to_write)} 条成功。")
        except Exception as e: print(f"⚠️ (后台)批量写入 '已通过' 失败，将重试: {e}")
approved_batcher = ApprovedBatcher()

def save_for_approved(job: Dict):
    info_row_map = {h: h for h in ["用户名", "小红书号", "主页网址", "邮箱", "搜索词", "个人简介", "粉丝数", "总点赞", "标记时间"]}
    info_row_map.update({"用户名": "username", "小红书号": "userid", "主页网址": "url", "个人简介": "bio", "粉丝数": "followers", "总点赞": "likes_total", "标记时间": "timestamp", "搜索词": "search_term"})
    info_data = {k: job.get(v, "") for k, v in info_row_map.items()}
    approved_batcher.add(info_data)

def save_for_review(job: Dict):
    row = [job.get("url", ""), job.get("userid", ""), job.get("username", ""), job.get("timestamp", "")]
    def writer(path):
        if not os.path.exists(path):
            wb = Workbook(); ws = wb.active; ws.append(REVIEW_COLS); wb.save(path)
        wb = load_workbook(path); wb.active.append(row); wb.save(path)
    safe_write_with_lock(MANUAL_REVIEW_EXCEL_PATH, writer)

def save_to_fetched_list(job: Dict):
    username, userid = job.get("username"), job.get("userid")
    def writer(path):
        sheet_name_user, sheet_name_id = "爬取用户名", "小红书号"
        if not os.path.exists(path):
            wb = Workbook(); ws_user = wb.active; ws_user.title = sheet_name_user; ws_user.append([sheet_name_user])
            ws_id = wb.create_sheet(sheet_name_id); ws_id.append([sheet_name_id]); wb.save(path)
        wb = load_workbook(path)
        if username: (wb[sheet_name_user] if sheet_name_user in wb.sheetnames else wb.create_sheet(sheet_name_user, 0)).append([username])
        if userid: (wb[sheet_name_id] if sheet_name_id in wb.sheetnames else wb.create_sheet(sheet_name_id, 1)).append([str(userid)])
        wb.save(path)
    safe_write_with_lock(FETCHED_USER_LIST_PATH, writer)

def save_for_training(job: Dict):
    info_row_map = {h: h for h in TRAINING_INFO_COLS}
    info_row_map.update({"用户名": "username", "小红书号": "userid", "主页网址": "url", "审核状态": "status", "AI预测概率": "ai_prob", "AI预测状态": "ai_decision", "个人简介": "bio", "粉丝数": "followers", "总点赞": "likes_total", "标记时间": "timestamp", "搜索词": "search_term"})
    info_row = {k: job.get(v, "") for k, v in info_row_map.items()}
    notes_rows = []; userid = job.get("userid") or job.get("username") or "unknown"
    futures = {i: img_pool.submit(download_and_convert_image, n.get("cover_url", ""), userid) for i, n in enumerate(job.get("notes", [])[:NUM_NOTES_TO_PROCESS]) if n.get("cover_url")}
    covers = {i: fut.result() for i, fut in futures.items()}
    for i, n in enumerate(job.get("notes", [])[:NUM_NOTES_TO_PROCESS]): notes_rows.append([job.get("userid", ""), i + 1, n.get("title", ""), parse_count_chinese(str(n.get("likes", "0"))), covers.get(i, ""), job.get("timestamp", "")])
    def writer(path):
        if not os.path.exists(path):
            wb = Workbook(); ws_info = wb.active; ws_info.title = INFO_SHEET; ws_info.append(TRAINING_INFO_COLS)
            ws_notes = wb.create_sheet(NOTES_SHEET); ws_notes.append(TRAINING_NOTES_COLS); wb.save(path)
        wb = load_workbook(path)
        ws_info = wb[INFO_SHEET] if INFO_SHEET in wb.sheetnames else wb.create_sheet(INFO_SHEET, 0)
        ws_info.append([info_row.get(h, "") for h in TRAINING_INFO_COLS])
        if notes_rows:
            ws_notes = wb[NOTES_SHEET] if NOTES_SHEET in wb.sheetnames else wb.create_sheet(NOTES_SHEET, 1)
            for r in notes_rows: ws_notes.append(r)
        wb.save(path)
    safe_write_with_lock(NEW_TRAINING_DATA_EXCEL, writer)

def process_job(job: Dict):
    job_id = job.get("job_id")
    if job_id and r.sismember(WAL_DONE_SET, job_id): return
    status = job.get("status")
    if status == "符合":
        try: save_for_approved(job)
        except Exception as e: print(f"❌ [后台] 保存 '已通过' 数据失败: {e}")
    if status == "人工审核":
        try: save_for_review(job); print(f"📋 (后台)已保存至待复审: {job.get('username') or job.get('userid')}")
        except Exception as e: print(f"❌ [后台] 保存 '待复审' 数据失败: {e}")
    try: save_to_fetched_list(job)
    except Exception as e: print(f"⚠️ [后台] 写入Excel去重列表失败: {e}")
    save_history_enabled = r.get(SAVE_HISTORY_ENABLED_KEY) == "1"
    if status == "符合" or (status == "不符合" and save_history_enabled):
        try: save_for_training(job)
        except Exception as e: print(f"⚠️ [后台] 写入训练数据失败: {e}")
    if job_id: r.sadd(WAL_DONE_SET, job_id)

def consume_jobs():
    print("✅ 后台消费线程已启动...")
    while not shutdown_event.is_set():
        try:
            job = job_q.get(timeout=1)
            process_job(job)
            job_q.task_done()
        except queue.Empty:
            approved_batcher.flush()
            continue
        except Exception as e: print(f"❌ 处理后台任务时发生严重错误: {e}"); time.sleep(2)

# ========== AI 预测核心函数 ==========
def run_ai_prediction(blogger_data: Dict) -> (float, str, str):
    if not AI_ENABLED_BY_FILE: return 0.0, "error", "AI模型未启用或加载失败"
    try:
        text_dim=TEXT_EMBEDDING_MODEL.get_sentence_embedding_dimension()or 768;image_dim=IMAGE_EMBEDDING_MODEL.get_sentence_embedding_dimension()or 512;bio=str(blogger_data.get('bio',''));notes_data=blogger_data.get('notes',[]);found_email=extract_and_normalize_email(bio);has_email=1 if found_email else 0;followers=parse_count_chinese(str(blogger_data.get('followers',0)));total_likes=parse_count_chinese(str(blogger_data.get('likes_total',0)));avg_likes_per_fan=total_likes/(followers+1)if followers>0 else 0;s_ratio,d_ratio=0,0
        if notes_data:likes_list=[parse_count_chinese(str(n.get('likes','0')))for n in notes_data[:NUM_NOTES_TO_PROCESS]];s_ratio=sum(1 for l in likes_list if l<10)/len(likes_list)if likes_list else 0;d_ratio=sum(1 for l in likes_list if 10<=l<100)/len(likes_list)if likes_list else 0
        bio_vec=TEXT_EMBEDDING_MODEL.encode([bio],convert_to_tensor=True,device=AI_DEVICE);titles=[n.get('title','')for n in notes_data[:NUM_NOTES_TO_PROCESS]];title_vec=TEXT_EMBEDDING_MODEL.encode(titles,convert_to_tensor=True,device=AI_DEVICE).mean(axis=0,keepdim=True)if titles else torch.zeros((1,text_dim),device=AI_DEVICE);image_urls=[note.get('cover_url')for note in notes_data if note.get('cover_url')];image_vec=torch.zeros((1,image_dim),device=AI_DEVICE)
        if image_urls:
            pil_images=[]
            for url in image_urls[:NUM_NOTES_TO_PROCESS]:
                if url:
                    try:pil_images.append(Image.open(io.BytesIO(http.get(url,timeout=5).content)).convert("RGB"))
                    except:pass
            if pil_images:image_vec=IMAGE_EMBEDDING_MODEL.encode(pil_images,convert_to_tensor=True,device=AI_DEVICE).mean(axis=0,keepdim=True)
        numeric_features=np.array([[has_email,followers,total_likes,avg_likes_per_fan,s_ratio,d_ratio]],dtype=np.float32);full_feature_np=np.concatenate([numeric_features,bio_vec.cpu().numpy(),title_vec.cpu().numpy(),image_vec.cpu().numpy()],axis=1)
        if full_feature_np.shape[1]<len(AI_SCALER.feature_names_in_):padded_arr=np.zeros((1,len(AI_SCALER.feature_names_in_)));padded_arr[:,:full_feature_np.shape[1]]=full_feature_np;full_feature_np=padded_arr
        scaled_features=AI_SCALER.transform(pd.DataFrame(full_feature_np,columns=AI_SCALER.feature_names_in_))
        with torch.no_grad():probability=AI_MODEL(torch.tensor(scaled_features,dtype=torch.float32).to(AI_DEVICE)).item()
        if probability>AI_ACCEPT_THRESHOLD:decision="符合";reason=f"模型概率 {probability:.2f} > {AI_ACCEPT_THRESHOLD}"
        elif probability>=AI_REJECT_THRESHOLD:decision="人工审核";reason=f"模型概率 {probability:.2f} 在 [{AI_REJECT_THRESHOLD}, {AI_ACCEPT_THRESHOLD}] 之间"
        else:decision="不符合";reason=f"模型概率 {probability:.2f} < {AI_REJECT_THRESHOLD}"
        return round(probability,4),decision,reason
    except Exception as e:
        print(f"❌ AI Prediction Error: {e}");return 0.0,"error",str(e)

def load_fetched_list_to_redis():
    if not os.path.exists(FETCHED_USER_LIST_PATH): print(f"ℹ️ 未找到 '{FETCHED_USER_LIST_PATH}'，跳过加载。"); return
    print(f"正在从 '{FETCHED_USER_LIST_PATH}' 加载已有用户到Redis...");
    try:
        wb=load_workbook(FETCHED_USER_LIST_PATH,read_only=True);pipe=r.pipeline();user_count=0;id_count=0
        if"爬取用户名"in wb.sheetnames:ws_user=wb["爬取用户名"];users_to_add=[row[0]for row in ws_user.iter_rows(min_row=2,values_only=True)if row and row[0] and row[0].strip()];pipe.sadd(USERNAMES_SET_KEY,*users_to_add);user_count=len(users_to_add)
        if"小红书号"in wb.sheetnames:ws_id=wb["小红书号"];ids_to_add=[str(row[0])for row in ws_id.iter_rows(min_row=2,values_only=True)if row and row[0] and str(row[0]).strip()];pipe.sadd(USERIDS_SET_KEY,*ids_to_add);id_count=len(ids_to_add)
        pipe.execute();print(f"✅ 从已有名单加载完成: {user_count}个用户名, {id_count}个小红书号。")
    except Exception as e:print(f"❌ 从 '{FETCHED_USER_LIST_PATH}' 加载失败: {e}")

# ========== API 路由 (核心修改处) ==========
@app.route("/usernames", methods=["GET"])
def get_usernames(): return jsonify(list(r.smembers(USERNAMES_SET_KEY)))
@app.route("/userids", methods=["GET"])
def get_userids(): return jsonify(list(r.smembers(USERIDS_SET_KEY)))

# ========== 【核心修复 I】 ==========
@app.route("/mark_data", methods=["POST"])
def mark_data():
    data = request.get_json(silent=True) or {}
    print(f"[{now_str()}] 收到 /mark_data 原始请求: user=({data.get('username')}|{data.get('userid')}), status={data.get('status')}")

    final_decision = (data.get("status") or "").strip()
    if not final_decision: 
        ai_prob, ai_decision, _ = run_ai_prediction(data)
        final_decision = ai_decision
        data['status'] = final_decision
        data['ai_prob'] = ai_prob
        data['ai_decision'] = ai_decision
    else:
        ai_prob, ai_decision, _ = run_ai_prediction(data)
        data['ai_prob'] = ai_prob
        data['ai_decision'] = ai_decision

    username, userid = (data.get("username") or "").strip(), (data.get("userid") or "").strip()
    if not final_decision or (not username and not userid): 
        return jsonify({"status": "error", "message": "缺少关键参数"}), 400
    
    with mark_data_lock:
        is_uname_member = r.sismember(USERNAMES_SET_KEY, username) if username else False
        is_uid_member = r.sismember(USERIDS_SET_KEY, str(userid)) if userid else False
        
        # 【逻辑变更】如果用户已在去重库中，并且新状态不是“人工审核”，则视为重复并拦截。
        # 这个修改允许一个之前被判为“不符合”的用户，可以被再次提交为“人工审核”状态。
        if (is_uname_member or is_uid_member) and final_decision != "人工审核":
            print(f"⚠️ 拦截到重复保存请求(已在Redis中): user=({username}|{userid})")
            return jsonify({"status": "duplicated"})
        
        # 【逻辑变更】只有当最终状态是“符合”或“不符合”时，才将用户加入去重库。
        # “人工审核”状态的用户不会被加入，以便将来还能被再次审核。
        if final_decision in ["符合", "不符合"]:
            if username: r.sadd(USERNAMES_SET_KEY, username)
            if userid: r.sadd(USERIDS_SET_KEY, str(userid))
            print(f"✅ (同步)用户 {username}|{userid} 已写入Redis去重集合。")

    if not data.get("email"): data['email'] = extract_and_normalize_email(data.get("bio", "")) or ""
    ts = now_str()
    job_id = f"{userid or username}:{int(time.time()*1000)}:{final_decision}"
    job = {**data, "job_id": job_id, "timestamp": ts, "status": final_decision}
    
    try:
        with open(WAL_FILE, "a", encoding="utf-8") as f: f.write(json.dumps(job, ensure_ascii=False) + "\n")
        job_q.put(job, timeout=0.1)
    except queue.Full: 
        print(f"❌ 服务繁忙，后台队列已满，无法处理: user=({username}|{userid})")
        return jsonify({"status": "error", "message": "服务繁忙，队列已满"}), 503
    except Exception as e: 
        print(f"❌ 写入WAL或入队时出错: {e}")
        return jsonify({"status": "error", "message": "内部错误"}), 500
    
    print(f"    -> 任务已入队: {final_decision} (AI预测: {ai_decision}, P={ai_prob:.4f}), user=({username}|{userid})")
    update_and_print_daily_stats(final_decision)
    return jsonify({"status": "ok"})

@app.route("/ai/settings", methods=["POST"])
def ai_settings(): enabled=bool(request.get_json(silent=True).get("enabled",False));r.set(AI_ENABLED_KEY,"1" if enabled else "0");print(f"⚙️ AI自动审核已 {'开启' if enabled else '关闭'}");return jsonify({"status":"ok","ai_enabled":enabled})
@app.route("/settings/save_history", methods=["POST"])
def save_history_settings(): enabled=bool(request.get_json(silent=True).get("enabled",False));r.set(SAVE_HISTORY_ENABLED_KEY,"1" if enabled else "0");print(f"⚙️ 保存“不符合”的训练数据已 {'开启' if enabled else '关闭'}");return jsonify({"status":"ok","save_history_enabled":enabled})
@app.route("/ai/decide", methods=["POST"])
def ai_decide(): data=request.get_json(silent=True);p,dec,why=run_ai_prediction(data);data['status']=dec;threading.Thread(target=mark_data_background,args=(data,)).start();return jsonify({"decision":dec,"reason":why,"p_base":p})
def mark_data_background(data): postJSON(f"http://127.0.0.1:{FLASK_PORT}/mark_data", data)
def postJSON(url,obj):requests.post(url,json=obj,timeout=5)
@app.route("/ai/suggest", methods=["POST"])
def ai_predict_only(): blogger_data = request.get_json(silent=True); p, dec, why = run_ai_prediction(blogger_data); return jsonify({"decision": dec, "reason": why, "p_base": p})

# ========== 【核心修复 II】 ==========
@app.route("/get_review_list", methods=["GET"])
def get_review_list():
    if not os.path.exists(MANUAL_REVIEW_EXCEL_PATH): 
        return jsonify([])
    try:
        def reader(path):
            wb = load_workbook(path)
            ws = wb.active
            if ws.max_row < 2:
                return [] # 文件存在但没有数据
            
            # 读取所有URL
            urls = [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True) if row and row[0]]
            
            # 【稳健性修改】删除数据行，而不是重建文件
            if urls:
                ws.delete_rows(2, len(urls))
                wb.save(path)
            
            return urls

        urls = safe_write_with_lock(MANUAL_REVIEW_EXCEL_PATH, reader) or []
        print(f"✅ 提供了 {len(urls)} 个待复审URL，并已清空列表。")
        return jsonify(urls)
    except Exception as e: 
        print(f"❌ 读取 '{MANUAL_REVIEW_EXCEL_PATH}' 失败: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500

def read_sheet_as_dicts(xlsx_path: str, sheet_name: str) -> List[Dict]:
    if not os.path.exists(xlsx_path): return []
    wb=load_workbook(xlsx_path,read_only=True);ws=wb[sheet_name]
    if ws.max_row<2:return[]
    rows=list(ws.iter_rows(values_only=True));headers=[str(h or "").strip()for h in rows[0]];return[dict(zip(headers,r))for r in rows[1:]]
def _filter_delta_rows(rows: List[Dict], min_dt: Optional[datetime]) -> List[Dict]:
    return [d for d in rows if(dt:=parse_dt(str(d.get("标记时间",""))))and dt>min_dt] if min_dt else rows
@app.route("/export_delta", methods=["GET"])
def export_delta():
    dataset=(request.args.get("dataset")or"approved").lower()
    curA_dt=parse_dt(r.get(CUR_APPROVED))if r.exists(CUR_APPROVED)else None
    wb=Workbook();wb.remove(wb.active)
    if dataset in("approved","both"):
        ws_info=wb.create_sheet("已通过-博主信息");ws_info.append(APPROVED_COLS)
        if os.path.exists(APPROVED_EXCEL_PATH):
            all_info=read_sheet_as_dicts(APPROVED_EXCEL_PATH,INFO_SHEET)
            for d in _filter_delta_rows(all_info,curA_dt):ws_info.append([d.get(k,"")for k in APPROVED_COLS])
    ts_name=datetime.now().strftime("%Y%m%d_%H%M%S");out_name=f"{DELTA_PREFIX}{ts_name}.xlsx";wb.save(out_name);nowS=now_str()
    if dataset in("approved","both"):r.set(CUR_APPROVED,nowS)
    print(f"📤 增量导出完成: {out_name}, 游标已更新至 {nowS}");return send_file(out_name,as_attachment=True,download_name=out_name)
@app.route("/rebuild_sets", methods=["POST"])
def rebuild_sets():
    try:
        print("🔁 开始从所有Excel文件重建Redis去重集合..."); r.delete(USERNAMES_SET_KEY); r.delete(USERIDS_SET_KEY)
        for path in [APPROVED_EXCEL_PATH, NEW_TRAINING_DATA_EXCEL, MANUAL_REVIEW_EXCEL_PATH]:
            if not os.path.exists(path): continue
            for row in read_sheet_as_dicts(path, INFO_SHEET if path != MANUAL_REVIEW_EXCEL_PATH else "Sheet"):
                if u := (row.get("用户名") or row.get("URL") or "").strip(): r.sadd(USERNAMES_SET_KEY, u)
                if i := (row.get("小红书号") or "").strip(): r.sadd(USERIDS_SET_KEY, str(i))
        load_fetched_list_to_redis()
        final_user_count,final_id_count=r.scard(USERNAMES_SET_KEY),r.scard(USERIDS_SET_KEY)
        print(f"✅ Redis去重集合已重建完成: {final_user_count} 用户名, {final_id_count} 小红书号。")
        return jsonify({"status":"ok","usernames":final_user_count,"userids":final_id_count})
    except Exception as e:return jsonify({"status":"error","msg":str(e)}),500
@app.route('/dashboard', methods=['GET'])
def dashboard_page():
    try:
        with open('dashboard.html','r',encoding='utf-8')as f:return render_template_string(f.read())
    except FileNotFoundError:return"Error: dashboard.html not found.",404
@app.route("/dashboard_stats", methods=["GET"])
def dashboard_stats():
    try:
        today_str=get_statistical_date_str();approved_today=int(r.get(f"daily_stats:approved:{today_str}")or 0);rejected_today=int(r.get(f"daily_stats:rejected:{today_str}")or 0);yesterday_dt=datetime.now()-timedelta(days=1);yesterday_str=get_statistical_date_str(yesterday_dt);approved_yesterday=int(r.get(f"daily_stats:approved:{yesterday_str}")or 0);rejected_yesterday=int(r.get(f"daily_stats:rejected:{yesterday_str}")or 0);pending_review_count=0
        if os.path.exists(MANUAL_REVIEW_EXCEL_PATH):
            def reader(path): return max(0,load_workbook(path).active.max_row-1)
            pending_review_count = safe_write_with_lock(MANUAL_REVIEW_EXCEL_PATH, reader) or 0
        return jsonify({"today":{"approved":approved_today,"rejected":rejected_today,"total":approved_today+rejected_today,},"yesterday":{"total":approved_yesterday+rejected_yesterday,},"pending_review":pending_review_count})
    except Exception as e:return jsonify({"status":"error","msg":str(e)}),500
@app.route('/open_folder', methods=['POST'])
def open_folder_route():
    key=(request.get_json()or{}).get('key');FILE_MAP={"approved":APPROVED_EXCEL_PATH,"review":MANUAL_REVIEW_EXCEL_PATH,"training":NEW_TRAINING_DATA_EXCEL,"output":os.getcwd()};file_path=FILE_MAP.get(key)
    if not file_path:return jsonify({"status":"error","message":"Invalid file key"}),400
    try:
        folder_path=os.path.dirname(os.path.abspath(file_path))if os.path.isfile(file_path)else os.path.abspath(file_path)
        if not os.path.exists(folder_path):os.makedirs(folder_path,exist_ok=True)
        system=platform.system()
        if system=="Windows":os.startfile(folder_path)
        elif system=="Darwin":subprocess.run(["open",folder_path])
        else:subprocess.run(["xdg-open",folder_path])
        print(f"📂 已请求打开文件夹: {folder_path}");return jsonify({"status":"ok","path":folder_path})
    except Exception as e:return jsonify({"status":"error","message":str(e)}),500
@app.route('/download_file', methods=['GET'])
def download_file_route():
    key=request.args.get('key');FILE_MAP={"approved":APPROVED_EXCEL_PATH,"review":MANUAL_REVIEW_EXCEL_PATH,"training":NEW_TRAINING_DATA_EXCEL};file_path=FILE_MAP.get(key)
    if not file_path:return"Invalid file key",400
    if not os.path.exists(file_path):
        if key == 'approved': wb=Workbook();ws=wb.active;ws.title=INFO_SHEET;ws.append(APPROVED_COLS);wb.save(file_path)
        elif key == 'review': wb=Workbook();ws=wb.active;ws.append(REVIEW_COLS);wb.save(file_path)
        elif key == 'training': wb=Workbook();ws_info=wb.active;ws_info.title=INFO_SHEET;ws_info.append(TRAINING_INFO_COLS);ws_notes=wb.create_sheet(NOTES_SHEET);ws_notes.append(TRAINING_NOTES_COLS);wb.save(file_path)
    return send_file(file_path,as_attachment=True)

# ========== 启动与退出 ==========
def graceful_shutdown(*args, **kwargs):
    print("\n⏹️ 正在优雅退出，请稍候...")
    shutdown_event.set()
    print("   - 等待任务队列清空..."); job_q.join()
    print("   - 正在强制刷新所有批处理数据到Excel..."); approved_batcher.flush(force=True)
    print("   - 刷新完成。")
    print("✅ 后台任务已安全处理完毕，服务退出。")
    os._exit(0)

if __name__ == "__main__":
    consumer_thread = threading.Thread(target=consume_jobs, daemon=True); consumer_thread.start()
    atexit.register(graceful_shutdown)
    signal.signal(signal.SIGINT, graceful_shutdown); signal.signal(signal.SIGTERM, graceful_shutdown)
    load_fetched_list_to_redis()
    print(f"🚀 [v5.5 - 关键修复版] 服务启动：http://127.0.0.1:{FLASK_PORT}")
    print(f"📊 数据面板请访问: http://localhost:{FLASK_PORT}/dashboard")
    from waitress import serve
    serve(app, host="0.0.0.0", port=FLASK_PORT, threads=16)


# === NEW: 同步补全用户名/ID 到 Redis，并异步补充到 "已爬取用户名.xlsx" ===
@app.route("/touch_user", methods=["POST"])
def touch_user():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    userid   = (data.get("userid") or "").strip()
    if not username and not userid:
        return jsonify({"status":"error","message":"missing username/userid"}), 400

    # 1) 同步：写入 Redis（立即生效，用于前端实时去重/高亮）
    with mark_data_lock:
        if username:
            r.sadd(USERNAMES_SET_KEY, username)
        if userid:
            r.sadd(USERIDS_SET_KEY, str(userid))
        print(f"✅ [/touch_user] 同步写入 Redis: ({username}|{userid})")

    # 2) 异步：Excel 写入仍走后台队列（可在空闲时批量落地）
    job = {
        "job_id": f"touch:{userid or username}:{int(time.time()*1000)}",
        "username": username,
        "userid": userid,
        "url": data.get("url",""),
        "timestamp": now_str(),
        "status": "touch"
    }
    try:
        with open(WAL_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(job, ensure_ascii=False) + "\\n")
        job_q.put(job, timeout=0.1)
    except Exception as e:
        print(f"⚠️ [/touch_user] 入队失败: {e}")

    return jsonify({"status":"ok"})
